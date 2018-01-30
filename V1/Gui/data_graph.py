from openpyxl import Workbook, load_workbook
import time
from tkinter.filedialog import askopenfilename, askdirectory
import tkinter as tk
from matplotlib.figure import Figure
import matplotlib.animation as animation
import threading


class DataGraph(object):
    def __init__(self, master, serial_port):
        self.master = master
        self.serial = serial_port
        #  See if I need a subplot to properly graph Live Data
        self.startbit = b's'
        self.start_time = 0
        self.stop_time = 0
        self.record = 0  # Can Start Recording Data
        self.threading_flag = 0  # Enable Threading
        self.graphing = 0   # Enable Graphing
        self.start_pressed = 0  # If the Start Button has been Pressed
        self.popup_enable = 0    # Popup Can be Shown
        self.sensors_switch = [0] * 24
        self.sensors_pressure = [[0 for x in range(1)] for y in range(1)]  # 2D Array of Sensor Pressure Values
        self.avail_sensors = []  # Initializing an Array of Possible Available Sensors
        self.selected_sensors = []  # Initializing an Array of Selected Sensors
        self.seconds = [0]  # Time Since Data has been Captured
        self.num_avail_sensors = 0  # Number of Available Sensors
        self.num_selected_sensors = 0   # Number of Selected Sensors
        self.calibration_filename = '_'
        self.sensitivity_filename = '_'
        self.destination_folder = '_'
        self.destination_filename = '_'
        self.calibration_values = [0] * 24  # Calibration values of the CDC value when no pressure Applied
        self.sensitivity_values = [0] * 24  # Sensitivity of Capacitive Sensor (Pa/pf)

        self.graph_seconds = [0]    # Subsection of Time since Data has been Captured
        self.graph_sensors_pressure = [[0 for x in range(1)] for y in range(1)]  # 2D Array of Subsection Sensor Pressure Values

        self.pyserial_thread = threading.Thread(target=self.thread_record_data, args=("do_run",))

        # self.cdc_pf = 3.8333E-4     # CDC Output code for pF change
        self.cdc_pf = 2.4414E-4   # CDC Output code to pF change

        self.live_plot_figure = Figure(figsize=(10, 5), dpi=100, tight_layout=True)
        self.sub_plot = self.live_plot_figure.add_subplot(1, 1, 1)

    def animate_graph(self, graph_frame):
        self.ani1 = animation.FuncAnimation(graph_frame, self.graph_data, interval=200, frames=60)
        #   frames*interval / 1000 (in seconds)

    def update_checkbox_sensors(self, s_toggle):
        self.selected_sensors = []     # Reinitializing Selected Sensors
        if self.sensors_switch[s_toggle-1] == 0:
            self.sensors_switch[s_toggle-1] = 1
        else:
            self.sensors_switch[s_toggle-1] = 0
        for i in range(0, len(self.sensors_switch)):
            if self.sensors_switch[i] == 1:
                self.selected_sensors.append(i+1)
        self.num_selected_sensors = len(self.selected_sensors)

    def record_start(self):
        self.serial.write_8('c')    # Setting the State to Continous Sensor Monitoring
        self.serial.write_8('v')    # Setting the mode to Continous Multiple Sensor Monitoring

    def reset_sensors(self):
        self.serial.flush_input()
        self.sensors_pressure = [[0 for x in range(1)] for y in range(self.num_selected_sensors)]
        self.graph_sensors_pressure = [[0 for x in range(1)] for y in range(self.num_selected_sensors)]
        self.serial.write_8('p')    # Setting the State to CDC Parameters
        self.serial.write_8('n')    # Setting the mode to Reset Sensors
        self.serial.write_dec(self.num_selected_sensors)     # Sending the number of Sensors to MCU
        for x in range(0, self.num_selected_sensors):
            self.serial.write_dec(self.selected_sensors[x])    # Writing Available sensors to the MCU

    def record_stop(self):
        self.serial.write_8('z')
        print("Serial Port Recording Closed")

    def button_start(self):
        if self.start_pressed == 0:
            self.threading_flag = 1
            self.record = 1
            self.reset_sensors()   # Reset the Number of Sensors and Enabled Sensors According to the Checkbox
            self.calibrate_glove()  # Calibrate the Glove by recording offset from Sensors
            self.record_start()  # Start Recording Sensor Data from Glove
            self.thread_graphing()  # Start Thread to continously Live Graph Data
            self.delay()    # Delay a second to Start getting some Data to graph
            self.graphing = 1
            self.start_pressed = 1
            print("Button Start")

    def button_stop(self):
        if self.start_pressed == 1:
            self.record = 0
            self.graphing = 0
            self.popup_enable = 1
            self.record_stop()
            self.thread_graphing()
            self.start_time = 0
            self.threading_flag = 0
            self.start_pressed = 0
            print("Button Stop")

    def ask_user_file(self, file_show):
        self.sensitivity_filename = askopenfilename()
        file_show.delete(0, tk.END)
        file_show.insert(0, self.sensitivity_filename)

    def ask_user_output_folder(self, folder_show):
        self.destination_filename = askdirectory()
        folder_show.delete(0, tk.END)
        folder_show.insert(0, self.destination_filename)

    def get_sensitivity(self):
        popup_sensitivity = tk.Toplevel()
        popup_sensitivity.grab_set()
        popup_sensitivity.geometry("1000x100")
        popup_sensitivity.wm_title("sensitivity Excel Data")
        label = tk.Label(popup_sensitivity, text="Select the sensitivity File For A Given Glove:")
        label.grid(row=0, column=0)

        b3 = tk.Button(popup_sensitivity, text="Browse:", command=lambda: self.ask_user_file(file_show))
        b3.grid(row=1, column=0)

        sensitivity_filename_string_var = tk.StringVar(popup_sensitivity, value=self.sensitivity_filename)
        file_show = tk.Entry(popup_sensitivity, textvariable=sensitivity_filename_string_var, width=150)
        file_show.grid(row=1, column=1)

        b1 = tk.Button(popup_sensitivity, text="Select Excel", command=lambda: [popup_sensitivity.destroy(), self.load_sensitivity(sensitivity_filename_string_var),
                                                                                popup_sensitivity.grab_release()])
        b1.grid(row=2, column=0)

        b2 = tk.Button(popup_sensitivity, text="Cancel", command=popup_sensitivity.destroy)
        b2.grid(row=2, column=1)

        popup_sensitivity.mainloop()

    def output_folder_popup(self):
        if self.popup_enable == 1:
            popup_output = tk.Toplevel()
            popup_output.grab_set()
            popup_output.geometry("1000x100")
            popup_output.wm_title("Select Output Folder to Save Data")
            label = tk.Label(popup_output, text="Output Folder:")
            label.grid(row=0, column=0)

            b4 = tk.Button(popup_output, text="Browse:", command=lambda:
                           self.ask_user_output_folder(folder_show))
            b4.grid(row=1, column=0)

            excel_folder_string_var = tk.StringVar(popup_output, value=self.destination_folder)
            folder_show = tk.Entry(popup_output, textvariable=excel_folder_string_var, width=150)
            folder_show.grid(row=1, column=1)

            label_file_name = tk.Label(popup_output, text="File Name:")
            label_file_name.grid(row=2, column=0)

            excel_file_string_var = tk.StringVar(popup_output, value=self.destination_filename)
            file_name_show = tk.Entry(popup_output, textvariable=excel_file_string_var, width=150)
            file_name_show.grid(row=2, column=1)

            b1 = tk.Button(popup_output, text="Save Excel File", command=lambda:
                           [self.save_data(str(excel_folder_string_var.get()) + '/' + str(excel_file_string_var.get()) + '.xlsx'), self.reset_data(), popup_output.destroy()])
            b1.grid(row=3, column=0)

            b2 = tk.Button(popup_output, text="Cancel", command=lambda: [self.reset_data(), popup_output.destroy()])
            b2.grid(row=3, column=1)

            self.popup_enable = 0
            popup_output.mainloop()

    def load_sensitivity(self, sensitivity_file_name):
        sensitivity_file = str(sensitivity_file_name.get())
        wb = load_workbook(sensitivity_file)
        ws_name = wb.get_sheet_names()[0]
        ws = wb.get_sheet_by_name(ws_name)

        for n in range(2, ws.max_row):
            if ws.cell(row=n, column=2).value != 0:
                self.sensitivity_values[n-2] = ws.cell(row=n, column=2).value
                self.avail_sensors.append(n-1)
        self.num_avail_sensors = len(self.avail_sensors)

    def calibrate_glove(self):
        for x in range(0, self.num_selected_sensors):
            self.serial.write_8('s')  # Setting State to Single
            self.serial.write_8('b')  # Setting mode to Single Sensor
            self.serial.write_dec(self.selected_sensors[x]-1)
            self.calibration_values[self.selected_sensors[x]-1] = self.serial.read_dec()
        print("calibration values:")
        print(self.calibration_values)
        print("Selected Sensors:")
        print(self.selected_sensors)
        print(self.num_selected_sensors)
        print("Available Sensors:")
        print(self.avail_sensors)
        print(self.num_avail_sensors)
        print("Sensitivity Values:")
        print(self.sensitivity_values)

    def save_data(self, output_excel):
        print("Saved")
        self.destination_filename = output_excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Sample #1"
        ws.cell(row=1, column=1, value="Time (s)")
        for k in range(0, self.num_selected_sensors):
            title = "Sensor " + str(k) + ": (Pa)"
            ws.cell(row=1, column=2+k, value=title)
        # Storing the values in rows
        for i in range(1, len(self.seconds)):
            ws.cell(row=i + 1, column=1, value=self.seconds[i-1])
            for j in range(0, self.num_selected_sensors):
                ws.cell(row=i + 1, column=2 + j, value=self.sensors_pressure[j][i-1])
        # Saving the Excel File
        wb.save(filename=self.destination_filename)

    def reset_data(self):
        self.sensors_pressure = [[0 for x in range(1)] for y in range(1)]  # Reinitializing Pressure from Sensor Array
        self.graph_sensors_pressure = [[0 for x in range(1)] for y in range(1)]  # Reinitializing Pressure from Sensor Array for Graphing
        self.graph_seconds = [0]  # Reinitializing Seconds Array for Graphing
        self.seconds = [0]  # Reinitializing Seconds Array
        print("data reset")

    def record_data(self):
        # start data collection
        startbit = b's'
        if self.record == 1:
            while self.serial.wait() == 0:
                pass
            while True:
                buffer = self.serial.read_8()
                print(buffer)
                if buffer == startbit:
                    for n in range(0, self.num_selected_sensors):
                        temp1 = self.serial.read_dec()
                        print(temp1)
                        print(n)
                        temp3 = (temp1 - self.calibration_values[self.avail_sensors[n]]) / (65535 - self.calibration_values[self.avail_sensors[n]-1]) * self.sensitivity_values[n]
                        self.sensors_pressure[n].append(temp3)  # Storing values in Larger Data Array
                    self.serial.flush_input()
                    break
                else:
                    self.serial.flush_input()

            count = time.time() - self.start_time
            self.seconds.append(count)

    def thread_record_data(self, arg):
        del arg
        # start data collection
        if self.record == 1:
            while getattr(self.pyserial_thread, "do_run", True):
                while self.serial.wait() == 0:
                    pass
                while True:
                    buffer = self.serial.read_8()
                    if buffer == self.startbit:
                        # print(buffer)
                        for n in range(0, self.num_selected_sensors):
                            temp_cdc = self.serial.read_dec()   # Reading the CDC Output from the AD7147 from all available sensors
                            print(str(n)+':'+str(temp_cdc))
                            # (CDC - Steady State) * (pF/CDC) * (Pa/pF)
                            temp_pressure = (temp_cdc - self.calibration_values[self.avail_sensors[n]-1]) * self.cdc_pf * self.sensitivity_values[self.avail_sensors[n]-1]
                            #   This makes all values less than the offset 0, This enables cleaner data to analyze
                            if temp_pressure < 0:
                                temp_pressure = 0
                            self.sensors_pressure[n].append(temp_pressure)  # Storing values in Larger Data Array
                        break   # This break exits the "while True" Loop

                count = time.time() - self.start_time
                self.seconds.append(count)

    def get_subdata(self):
        for n in range(0, self.num_selected_sensors):
            self.graph_sensors_pressure[n].append(self.sensors_pressure[n][-1])
        self.graph_seconds.append(self.seconds[-1])

    def update_single_fig(self):
        self.sub_plot.clear()
        self.sub_plot.set_title("Individual Sensor Readout")  # Setting the Title of the Graph
        self.sub_plot.set_ylabel("Pressure (Pa)")  # Set ylabels to Pressure in Psi
        self.sub_plot.set_xlabel("Time (s)")  # Set xlabel to Time in Seconds

        # only show 20 seconds worth of data
        if self.seconds[-1] < 20:
            self.sub_plot.set_xlim([0, 20])
        else:
            time_low = self.graph_seconds[-1] - 20
            self.sub_plot.set_xlim([time_low, self.graph_seconds[-1]])

        # Setting the Y limit to show PSI from 0 - 5 psi
        self.sub_plot.set_ylim([0, 20])
        self.sub_plot.ticklabel_format(useOffset=False)

    def init_single_fig(self):
        self.sub_plot.clear()
        self.sub_plot.set_title("Individual Sensor Readout")  # Setting the Title of the Graph
        self.sub_plot.set_ylabel("Pressure (Psi)")  # Set ylabels to Pressure in Psi
        self.sub_plot.set_xlabel("Time (s)")  # Set xlabel to Time in Seconds

        # Setting the Y limit to show PSI from 0 - 5 psi
        self.sub_plot.set_ylim([0, 20])
        self.sub_plot.ticklabel_format(useOffset=False)

        # sub_fig = self.live_plot_figure.set()

    def graph_data(self, i):
        del i
        if self.graphing == 1:
            self.get_subdata()
            self.update_single_fig()
            for n in range(0, self.num_selected_sensors):
                self.sub_plot.plot(self.graph_seconds, self.graph_sensors_pressure[n])

    def thread_graphing(self):
        if self.threading_flag == 1:
            if self.pyserial_thread.is_alive():
                self.pyserial_thread.do_run = False
                self.pyserial_thread.join()
                print('Thread Stopped')
            else:
                self.pyserial_thread = threading.Thread(target=self.thread_record_data, args=("do_run",))
                self.pyserial_thread.start()
                self.start_time = time.time()
                print("Starting Thread")

    @staticmethod
    def delay():
        time.sleep(0.01)

    def update_total_fig(self, arg):
        del arg
        while getattr(self.pyserial_thread, "do_run", True):
            time.sleep(0.001)
            print(self.avail_sensors)

    @staticmethod
    def quit_gui():
        quit()

    @staticmethod
    def print_name():
        print("Carl!")
