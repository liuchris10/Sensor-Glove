import tkinter as tk
from tkinter import ttk
from tkinter import *
import matplotlib
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2TkAgg
from matplotlib.figure import Figure
import matplotlib.animation as animation
from matplotlib import style
import serial_1 as sp1
from openpyxl import Workbook, load_workbook
import time
from tkinter.filedialog import askopenfilename


LARGE_FONT = ("Verdana", 12)
TITLE_FONT = ("Times", 16)
NORM_FONT = ("Helvetica", 10)
SMALL_FONT = ("Helvetica", 8)
style.use("ggplot")
matplotlib.use("TkAgg")

f = Figure(figsize=(5, 5), dpi=100, tight_layout=True)
a = f.add_subplot(1, 1, 1)


start_time = 0
stop_time = 0
record = 0
sensors_switch = [0]*24
sensors_pressure = [[0 for x in range(1)] for y in range(1)]     # 2D Array of Sensor Pressure Values
avail_sensors = []  # Initializing an Array of Possible Available Sensors
seconds = [0]    # Time Since Data has been Captured
num_sensors = 0
calibration_filename = '_'
calibration_values = [0]*24     # Calibration values of the CDC value when no pressure Applied
sensitivity_values = [0]*24     # Sensitivity of Capacitive Sensor (Pa/pf)

def init_gui():
    app = SensorGlove()
    ani1 = animation.FuncAnimation(f, recorddata, interval=100)
    app.geometry("1280x720")
    app.mainloop()


def popupmsg(msg):
    popup = tk.Tk()
    popup.wm_title("Pop-Up Message!")
    label = ttk.Label(popup, text=msg, font=NORM_FONT)
    label.grid(row=0, column=0)
    b1 = ttk.Button(popup, text="Okay", command=popup.destroy)
    b1.grid(row=0, column=1)
    popup.mainloop()


def record_start(value):
    global record
    global start_time
    global num_sensors
    global avail_sensors
    global sensors_pressure

    if value == 0:
        sp1.flush_input()
        sensors_pressure = [[0 for x in range(1)] for y in range(num_sensors)]
        sp1.write_8('p')
        sp1.write_8('n')
        sp1.write_dec(num_sensors)
        for x in range(0, num_sensors):
            sp1.write_dec(avail_sensors[x])
        sp1.write_8('c')
        sp1.write_8('v')
        record = 1
        start_time = time.time()


def record_stop(value):
    global record
    global start_time
    global seconds
    global sensors_pressure
    global num_sensors
    if value == 1:
        record = 0
        sp1.write_8('z')
        start_time = 0
        save_data()  # Saving the Data to an Excel Spread Sheet
    sensors_pressure = [[0 for x in range(1)] for y in range(num_sensors)]  # Reinitializing Pressure from Sensor Array
    seconds = [0]    # Reinitializing Seconds Array


def quit_gui(self):
    self.quit()


def print_name():
        print("Carl!")

def ask_user_file(file_show):
    global sensitivity_filename
    sensitivity_filename = askopenfilename()
    file_show.delete(0, END)
    file_show.insert(0, sensitivity_filename)

def get_sensitivity():
    global sensitivity_filename
    popup_sensitivity = tk.Toplevel()
    popup_sensitivity.grab_set()
    popup_sensitivity.geometry("1000x100")
    popup_sensitivity.wm_title("sensitivity Excel Data")
    label = ttk.Label(popup_sensitivity, text="Select the sensitivity File For A Given Glove:", font=NORM_FONT)
    label.grid(row=0, column=0)

    b3 = ttk.Button(popup_sensitivity, text="Browse:", command=lambda: ask_user_file(file_show))
    b3.grid(row=1, column=0)

    sensitivity_filename_string_var = StringVar(popup_sensitivity, value = sensitivity_filename)
    file_show = ttk.Entry(popup_sensitivity, font = NORM_FONT, textvariable = sensitivity_filename_string_var, width = 150)
    file_show.grid(row=1, column=1)


    b1 = ttk.Button(popup_sensitivity, text="Select Excel", command=lambda:[load_sensitivity(sensitivity_filename_string_var), popup_sensitivity.grab_release(), popup_sensitivity.destroy()])
    b1.grid(row=2, column=0)

    b2 = ttk.Button(popup_sensitivity, text="Cancel", command=popup_sensitivity.destroy)
    b2.grid(row=2, column=1)

    popup_sensitivity.mainloop()

def load_sensitivity(sensitivity_file_name):
    global avail_sensors
    global sensitivity_values
    sensitivity_file = str(sensitivity_file_name.get())
    wb = load_workbook(sensitivity_file)
    ws_name = wb.get_sheet_names()[0]
    ws = wb.get_sheet_by_name(ws_name)

    for n in range(2, ws.max_row):
        if ws.cell(row=n, column=2).value != 0:
            sensitivity_values[n-2] = ws.cell(row=n, column=2).value
            avail_sensors.append(n-1)


def get_calibration():



def calibrate_glove():
    get_sensitivity()
    get_calibration()

def save_data():
    global seconds
    global num_sensors
    global sensors_pressure
    dest_filename = 'C:/Users/Intern/Documents/Data/test2.xlsx'
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample #1"
    ws.cell(row=1, column=1, value="Time (s)")
    for k in range(0, num_sensors):
        title = "Sensor " + str(k) + ": (Pa)"
        ws.cell(row=1, column=2+k, value=title)
    # Storing the values in rows
    for i in range(1, len(seconds)):
        ws.cell(row=i + 1, column=1, value=seconds[i-1])
        for j in range(0, num_sensors):
            ws.cell(row=i + 1, column=2 + j, value=sensors_pressure[j][i-1])
    # Saving the Excel File
    wb.save(filename=dest_filename)


def update_single_fig():
    global seconds
    a.clear()
    a.set_title("Individual Sensor Readout")  # Setting the Title of the Graph
    a.set_ylabel("Pressure (Psi)")  # Set ylabels to Pressure in Psi
    a.set_xlabel("Time (s)")  # Set xlabel to Time in Seconds

    # only show 20 seconds worth of data
    if seconds[-1] < 20:
        a.set_xlim([0, 20])
    else:
        time_low = seconds[-1] - 20
        a.set_xlim([time_low, seconds[-1]])

    # Setting the Y limit to show PSI from 0 - 5 psi
    a.set_ylim([0, 5])
    a.ticklabel_format(useOffset=False)


def update_total_fig():
    print('Hello')


def recorddata(i):
    # start data collection
    global record
    global sensors_pressure
    global seconds
    startbit = b's'
    if record == 1:
        while sp1.wait() == 0:
            pass
        # sp1.flush_input()
        while True:
            buffer = sp1.read_8()
            if (buffer == startbit):
                for n in range(0, num_sensors):
                    temp1 = sp1.read_dec()
                    print(temp1)
                    temp3 = (temp1 - 40000) / (65536 - 40000) * 5
                    sensors_pressure[n].append(temp3)  # Storing values in Larger Data Array
                sp1.flush_input()
                break
            else:
                sp1.flush_input()

        count = time.time() - start_time
        seconds.append(count)

        update_single_fig()
        a.plot(seconds, sensors_pressure[0]) # seconds, sensors_pressure[1], seconds, sensors_pressure[2])
        # a.plot(seconds, sensors_pressure)


class SensorGlove(tk.Tk):
    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        tk.Tk.iconbitmap(self, default="sw-caution.gif")
        tk.Tk.wm_title(self, "Sensor Glove Gui")

        container = tk.Frame(self)
        container.grid(row=0, column=0)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.create_menu(container)
        self.frames = {}

        for F in (IntroPage, HelpPage, MainPage):
            frame = F(container, self)
            self.frames[F] = frame
            frame.grid(row=0, column=0, sticky="nsew")

        self.show_frame(IntroPage)

    def show_frame(self, cont):
        frame = self.frames[cont]
        frame.tkraise()

    def create_menu(self, container):
        menubar = tk.Menu(container)

        filemenu = tk.Menu(menubar, tearoff=0)
        filemenu.add_command(label="Save Graph", command=lambda: popupmsg("Not supported just yet!"))
        filemenu.add_command(label="Export Excel", command=lambda: popupmsg("Not supported just yet!"))
        filemenu.add_separator()
        filemenu.add_command(label="Exit", command=quit)
        menubar.add_cascade(label="File", menu=filemenu)

        submenu = tk.Menu(menubar, tearoff=1)
        submenu.add_command(label="Sensor Calibration", command=lambda: popupmsg("Not supported just yet!"))
        submenu.add_command(label="Save settings", command=lambda: popupmsg("Not supported just yet!"))
        menubar.add_cascade(label="Settings", menu=submenu)

        editmenu = tk.Menu(menubar, tearoff=1)
        editmenu.add_command(label="Help Menu", command=lambda: self.show_frame(HelpPage))
        editmenu.add_separator()
        editmenu.add_command(label="Print Name", command=lambda: print_name())
        menubar.add_cascade(label="Help", menu=editmenu)

        tk.Tk.config(self, menu=menubar)


class IntroPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        label1 = tk.Label(self, text="Welcome to Sensor Glove Gui!", font=LARGE_FONT)
        label1.grid(row=0, column=0)

        label2 = tk.Label(self, text="Designed by the Flexible Electronics Lab @ UCSD", font=TITLE_FONT)
        label2.grid(row=1, column=0)

        button1 = ttk.Button(self, text="Start!", command=lambda: controller.show_frame(MainPage))
        button1.grid(row=2, column=0)

        button2 = ttk.Button(self, text="Quit", command=lambda: quit())
        button2.grid(row=3, column=0)

        button3 = ttk.Button(self, text="Help Me!", command=lambda: controller.show_frame(HelpPage))
        button3.grid(row=4, column=0)

        button4 = ttk.Button(self, text="Calibrate", command=lambda: calibrate_glove())
        button4.grid(row=5, column=0)


class HelpPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)

        label = tk.Label(self, text="Welcome to Help Menu!", font=TITLE_FONT)
        label.grid(row=0, column=0)

        button1 = tk.Button(self, text="go Back to intro Page", command=lambda: controller.show_frame(IntroPage))
        button1.grid(row=1, column=0)



class MainPage(tk.Frame):
    def __init__(self, parent, controller):
        tk.Frame.__init__(self, parent)
        self.parent = parent
        self.container = controller
        self.create_checkbox()
        self.create_button()

        canvas_frame = tk.Frame(self)
        canvas = FigureCanvasTkAgg(f, canvas_frame)
        canvas.show()
        canvas.get_tk_widget().pack()
        canvas_frame.grid(row=6, column=0, columnspan=5)

        toolbar_frame = tk.Frame(self)
        toolbar = NavigationToolbar2TkAgg(canvas, toolbar_frame)
        toolbar.update()
        canvas._tkcanvas.pack()
        toolbar_frame.grid(row=7, column=0, columnspan=2)

    def update_checkbox_sensors(self, s_toggle):
        global avail_sensors
        global sensors_switch
        avail_sensors = []
        global num_sensors
        if sensors_switch[s_toggle-1] == 0:
            sensors_switch[s_toggle-1] = 1
        else:
            sensors_switch[s_toggle-1] = 0
        for i in range(0,len(sensors_switch)):
            if sensors_switch[i] == 1:
                avail_sensors.append(i+1)
        num_sensors = len(avail_sensors)

    def create_checkbox(self):
        button = []
        palm = tk.Frame(self)
        palm.grid(row=0, column=0, columnspan=4)
        fingers = tk.Frame(self)
        fingers.grid(row=0, column=5, columnspan=4)
        palm_title = tk.Label(palm, text='Palm', font=TITLE_FONT)
        palm_title.grid(row=0)
        right_title = tk.Label(fingers, text='Fingers', font=TITLE_FONT)
        right_title.grid(row=0)
        for n in range(1, 25):
            checkbox_name = "Sensor %d" % n
            if n > 12:
                button.append(tk.Checkbutton(fingers, text=checkbox_name, onvalue=1, offvalue=0, command=lambda n=n: self.update_checkbox_sensors(n)))
                if n > 20:
                    row = 3
                    column = 5 + (n-21)
                elif n > 16:
                    row = 2
                    column = 5 + (n-17)
                else:
                    row = 1
                    column = 5 + (n-13)
            else:
                button.append(tk.Checkbutton(palm, text=checkbox_name, onvalue=1, offvalue=0, command=lambda n=n: self.update_checkbox_sensors(n)))
                if n > 8:
                    row = 3
                    column = 0 + (n-9)
                elif n > 4:
                    row = 2
                    column = 0 + (n-5)
                else:
                    row = 1
                    column = 0 + (n-1)
            button[n-1].grid(row=row, column=column)

    def create_button(self):
        button_frame = tk.Frame(self)
        button2 = ttk.Button(button_frame, text="Start", command=lambda: record_start(record))
        button2.grid(row=0, column=0)
        button3 = ttk.Button(button_frame, text="Stop", command=lambda: record_stop(record))
        button3.grid(row=0, column=1)
        button_frame.grid(row=4, column=0)


def main():
    init_gui()


if __name__ == "__main__":
    main()
